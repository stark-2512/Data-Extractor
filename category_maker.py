from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np

def make(raw_category_table, type):
    # print(len(raw_category_table))
    
    df_category = pd.DataFrame(raw_category_table)
    
    # Drop useless columns
    col_to_drop = [4,5,7,8,11,12,14,15,18,19,21,22,25,26,28,29]
    df_category = df_category.drop(col_to_drop, axis=1)
    
    
    # bring total columns to front
    total_col = [23, 24, 27]
    all_col = df_category.columns.to_list()
    string_col = [0, 1]

    # Bringing last 3 columns in front
    modified_col = total_col + [col for col in all_col if col not in total_col]
    df_category = df_category[modified_col]
    df_category.drop(0, inplace=True)

    # filtering useless data
    df_category = df_category.replace('.', np.nan)
    df_category = df_category.replace('-', np.nan)
    
    # convert string to float
    col_contain_str = [2, 3, 9, 10, 16, 17, 23, 24]
    col_contain_str_with_percent = [6, 13, 20, 27]

    for col1 in col_contain_str:
        df_category[col1] = df_category[col1].str.replace(',', '').astype(float)

    for col1 in col_contain_str_with_percent:
        df_category[col1] = df_category[col1].str.replace('%', '').astype(float)
    
    # calculating growth by percentage
    
    total_percent = []

    maxi = df_category[23].max()
    for row in df_category[23]:
        per = (row/maxi) * 100
        total_percent.append(per)

    bpcl_percent = []

    maxi = df_category[2].max()
    for row in df_category[2]:
        per = (row/maxi) * 100
        bpcl_percent.append(per)
    
    # writing these two percent in the dataframe
    df_category.insert(3, 'Total Percent', total_percent)
    df_category.insert(6, 'BPCL Percent', bpcl_percent)    

    # sorting
    df_category = df_category.sort_values(['Total Percent'], ascending=False)
    
    
    # Giving Columns their real names
    df_category = df_category.rename(columns={
        0 : 'Region',
        1 : 'State',
        2 : 'BPC CY',
        3 : 'BPC LY',
        6 : 'BPC Growth',
        9 : 'IOC CY',
        10 : 'IOC LY',
        13 : 'IOC Growth',
        16 : 'HPC CY',
        17 : 'HPC LY',
        20 : 'HPC Growth',
        23 : 'Total CY',
        24 : 'Total LY',
        27 : 'Total Growth',
    })

    # Making Market Share Columns
    companies = ["BPC", "HPC", "IOC"]
    # 1. For BPCL
    
    for company in companies:
        df_company_msh = pd.DataFrame(columns=[f"{company} MSH CY", f"{company} MSH LY", f"{company} Diffrence"])
        i = 0
        for _,row in df_category.iterrows():
            cy = (row[f"{company} CY"] / row["Total CY"]) * 100
            ly = (row[f"{company} LY"] / row["Total LY"]) * 100
            new_row_for_msh = {
                f"{company} MSH CY" : cy,
                f"{company} MSH LY" : ly,
                f"{company} Diffrence": cy-ly
            }
            
            df_company_msh = pd.concat([df_company_msh, pd.DataFrame(new_row_for_msh, index=[i])], ignore_index=True)
            i += 1
        
        cols_to_insert = [f"{company} MSH CY", f"{company} MSH LY", f"{company} Diffrence"]
        for col in cols_to_insert:
            num_row, num_col =  df_category.shape
            df_category.insert(num_col, col, df_company_msh[col])
        # print(df_company_msh)

    

    # writing in excel
    df_category.to_excel(f'{type}.xlsx', engine='openpyxl', index=False)


    # coloring
    workbook = load_workbook(f'{type}.xlsx')
    sheet = workbook['Sheet1']
    green = '05B322'
    red = 'E41606'
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        value1 = row[3].value
        value2 = row[6].value
        if value1 != None and value2 != None:
            if value1 < value2:
                row[6].fill = PatternFill(start_color=green, end_color=green, fill_type='solid')
                row[3].fill = PatternFill(start_color=red, end_color=red, fill_type='solid')
            else:
                row[6].fill = PatternFill(start_color=red, end_color=red, fill_type='solid')
                row[3].fill = PatternFill(start_color=green, end_color=green, fill_type='solid')

    workbook.save(f'{type}.xlsx')